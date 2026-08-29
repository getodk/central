"""API tests against a stubbed central-backend.

The point of most of these is the permission boundary: Studio must never show
or change anything the caller's own Central account cannot reach.
"""

import csv
import io
import zipfile

import pytest
from conftest import FIXTURES
from fastapi.testclient import TestClient

from studio import central, db
from studio.central import CentralError
from studio.main import app

ALICE = {"id": 7, "displayName": "Alice", "email": "alice@example.org"}
BOB = {"id": 9, "displayName": "Bob", "email": "bob@example.org"}

# Alice can see both projects; Bob only project 2.
USERS = {"alice-token": ALICE, "bob-token": BOB}
VISIBLE = {"alice-token": [1, 2], "bob-token": [2]}
PROJECTS = {1: {"id": 1, "name": "Statistics"}, 2: {"id": 2, "name": "Health"}}

ROOT_CSV = (
    "SubmissionDate,hh-village,hh-hhsize,meta-instanceID,KEY\n"
    "2026-01-31T04:15:00.000Z,Nadi,4,uuid:a,uuid:a\n"
)
MEMBER_CSV = "name,sex,age,langs,PARENT_KEY,KEY\nAna,2,34,en fj,uuid:a,uuid:a/member[1]\n"


@pytest.fixture(autouse=True)
def stub_central(monkeypatch):
    published = {}

    def current_user(self):
        user = USERS.get(self.token)
        if user is None:
            raise CentralError(401, "expired")
        return user

    def projects(self):
        current_user(self)
        return [PROJECTS[pid] for pid in VISIBLE[self.token]]

    def forms(self, project_id):
        return [{"xmlFormId": "household_survey", "name": "Household Survey",
                 "version": "1", "state": "open", "submissions": 1}]

    def form_xml(self, project_id, xml_form_id, version=None):
        return (FIXTURES / "household.xml").read_text()

    def submissions_csv_zip(self, project_id, xml_form_id, **kwargs):
        return {"household_survey.csv": ROOT_CSV, "household_survey-member.csv": MEMBER_CSV}

    def submission_count(self, project_id, xml_form_id):
        return 1

    def create_form(self, project_id, xlsx, form_id_fallback):
        if form_id_fallback in published:
            raise CentralError(409, "A resource already exists with this id")
        published[form_id_fallback] = xlsx
        return {"xmlFormId": form_id_fallback}

    def create_draft(self, project_id, xml_form_id, xlsx, form_id_fallback):
        published[xml_form_id] = xlsx
        return {"success": True}

    for name, fn in [
        ("current_user", current_user), ("projects", projects), ("forms", forms),
        ("form_xml", form_xml), ("submissions_csv_zip", submissions_csv_zip),
        ("submission_count", submission_count), ("create_form", create_form),
        ("create_draft", create_draft),
    ]:
        monkeypatch.setattr(central.Client, name, fn)

    monkeypatch.setattr(
        central.Client, "login",
        staticmethod(lambda email, password: {"token": "alice-token"}
                     if password == "correct" else _reject()),
    )

    # Start every test from an empty store so they stay independent.
    db.init()
    with db.connect() as conn:
        conn.execute("DELETE FROM questionnaire_versions")
        conn.execute("DELETE FROM questionnaires")
    yield


def _reject():
    raise CentralError(401, "bad credentials")


@pytest.fixture
def client():
    with TestClient(app) as test_client:
        yield test_client


def auth(token="alice-token"):
    return {"Authorization": f"Bearer {token}"}


def new_document(form_id="hh"):
    return {
        "title": "Household", "formId": form_id, "version": "",
        "languages": ["English (en)"], "defaultLanguage": "English (en)",
        "choiceLists": [{"name": "yn", "options": [
            {"value": "1", "label": {"English (en)": "Yes"}},
            {"value": "0", "label": {"English (en)": "No"}}]}],
        "items": [
            {"kind": "question", "type": "text", "name": "village",
             "label": {"English (en)": "Village"}, "required": True},
            {"kind": "question", "type": "select_one", "name": "owns",
             "label": {"English (en)": "Owns home"}, "choiceList": "yn"},
        ],
    }


# -- authentication --------------------------------------------------------


def test_the_page_and_health_check_answer_head_requests(client):
    # nginx caching tests and Docker health checks both use HEAD.
    assert client.head("/studio/", follow_redirects=False).status_code == 200
    assert client.head("/studio", follow_redirects=False).status_code == 308
    assert client.head("/studio/api/health").status_code == 200
    assert client.head("/studio/static/app.js").status_code == 200


def test_sign_in_returns_a_token(client):
    response = client.post("/studio/api/session",
                           json={"email": "alice@example.org", "password": "correct"})
    assert response.status_code == 200
    assert response.json()["token"] == "alice-token"


def test_bad_credentials_are_rejected(client):
    response = client.post("/studio/api/session",
                           json={"email": "alice@example.org", "password": "wrong"})
    assert response.status_code == 401


def test_cookie_adoption_needs_the_client_header_and_a_cookie(client):
    # Without the custom header a cross-site page could trigger this endpoint.
    assert client.post("/studio/api/session/from-cookie").status_code == 400
    # With the header but no Central cookie there is simply nothing to adopt.
    assert client.post("/studio/api/session/from-cookie",
                       headers={"X-Studio-Client": "1"}).status_code == 204
    # A valid Central cookie is exchanged for a bearer token.
    client.cookies.set("__Host-session", "alice-token")
    response = client.post("/studio/api/session/from-cookie", headers={"X-Studio-Client": "1"})
    assert response.status_code == 200
    assert response.json()["user"]["displayName"] == "Alice"
    client.cookies.clear()


def test_endpoints_require_a_token(client):
    assert client.get("/studio/api/projects").status_code == 401
    assert client.get("/studio/api/projects", headers=auth("stale")).status_code == 401


def test_projects_are_limited_to_the_callers_own(client):
    assert [p["id"] for p in client.get("/studio/api/projects", headers=auth()).json()] == [1, 2]
    assert [p["id"] for p in client.get("/studio/api/projects", headers=auth("bob-token")).json()] == [2]


# -- questionnaires --------------------------------------------------------


def test_create_read_update_delete(client):
    created = client.post("/studio/api/questionnaires", headers=auth(),
                          json={"projectId": 1, "document": new_document()})
    assert created.status_code == 201
    record = created.json()
    assert record["createdBy"] == "Alice"

    listed = client.get("/studio/api/questionnaires?projectId=1", headers=auth()).json()
    assert [r["id"] for r in listed] == [record["id"]]

    document = new_document()
    document["title"] = "Renamed"
    saved = client.put(f"/studio/api/questionnaires/{record['id']}", headers=auth(), json=document)
    assert saved.json()["title"] == "Renamed"

    assert client.delete(f"/studio/api/questionnaires/{record['id']}", headers=auth()).status_code == 204
    assert client.get(f"/studio/api/questionnaires/{record['id']}", headers=auth()).status_code == 404


def test_a_questionnaire_is_invisible_outside_its_project(client):
    record = client.post("/studio/api/questionnaires", headers=auth(),
                         json={"projectId": 1, "document": new_document()}).json()

    # Bob has no access to project 1.
    assert client.get(f"/studio/api/questionnaires/{record['id']}",
                      headers=auth("bob-token")).status_code == 403
    assert client.put(f"/studio/api/questionnaires/{record['id']}",
                      headers=auth("bob-token"), json=new_document()).status_code == 403
    assert client.delete(f"/studio/api/questionnaires/{record['id']}",
                         headers=auth("bob-token")).status_code == 403
    assert client.get("/studio/api/questionnaires?projectId=1",
                      headers=auth("bob-token")).status_code == 403


def test_editing_keeps_a_version_history(client):
    record = client.post("/studio/api/questionnaires", headers=auth(),
                         json={"projectId": 1, "document": new_document()}).json()
    document = new_document()
    document["title"] = "Second"
    client.put(f"/studio/api/questionnaires/{record['id']}", headers=auth(), json=document)

    versions = client.get(f"/studio/api/questionnaires/{record['id']}/versions", headers=auth()).json()
    assert len(versions) == 2

    restored = client.post(
        f"/studio/api/questionnaires/{record['id']}/versions/{versions[-1]['id']}/restore",
        headers=auth(),
    ).json()
    assert restored["document"]["title"] == "Household"


# -- xlsform and publishing ------------------------------------------------


def test_xlsform_download(client):
    response = client.post("/studio/api/xlsform", json=new_document())
    assert response.status_code == 200
    assert "hh.xlsx" in response.headers["content-disposition"]
    from openpyxl import load_workbook
    workbook = load_workbook(io.BytesIO(response.content))
    assert set(workbook.sheetnames) == {"survey", "choices", "settings"}


def test_xlsform_refuses_an_invalid_questionnaire(client):
    document = new_document()
    document["items"][0]["name"] = "not valid"
    response = client.post("/studio/api/xlsform", json=document)
    assert response.status_code == 400


def test_publish_creates_a_form_then_reports_the_clash(client):
    record = client.post("/studio/api/questionnaires", headers=auth(),
                         json={"projectId": 1, "document": new_document()}).json()

    first = client.post(f"/studio/api/questionnaires/{record['id']}/publish",
                        headers=auth(), json={"document": new_document(), "mode": "new"})
    assert first.status_code == 200
    assert first.json()["xmlFormId"] == "hh"
    assert first.json()["version"].isdigit()

    reloaded = client.get(f"/studio/api/questionnaires/{record['id']}", headers=auth()).json()
    assert reloaded["publishedAs"] == "hh"

    again = client.post(f"/studio/api/questionnaires/{record['id']}/publish",
                        headers=auth(), json={"document": new_document(), "mode": "new"})
    assert again.status_code == 409
    assert "new draft" in again.json()["detail"]

    draft = client.post(f"/studio/api/questionnaires/{record['id']}/publish",
                        headers=auth(), json={"document": new_document(), "mode": "draft"})
    assert draft.status_code == 200


def test_publish_is_blocked_for_other_projects(client):
    record = client.post("/studio/api/questionnaires", headers=auth(),
                         json={"projectId": 1, "document": new_document()}).json()
    response = client.post(f"/studio/api/questionnaires/{record['id']}/publish",
                           headers=auth("bob-token"), json={"document": new_document()})
    assert response.status_code == 403


def test_import_round_trips_into_a_questionnaire(client):
    from studio import xlsform
    from studio.models import Questionnaire

    data = xlsform.to_workbook(Questionnaire(**new_document()))
    response = client.post("/studio/api/import", headers=auth(),
                           files={"file": ("form.xlsx", data,
                                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert response.status_code == 200
    assert response.json()["document"]["formId"] == "hh"
    assert response.json()["warnings"] == []


# -- export ----------------------------------------------------------------


def test_export_produces_a_zip_of_labelled_files(client):
    response = client.post("/studio/api/export", headers=auth(), json={
        "projectId": 1, "xmlFormId": "household_survey", "formats": ["stata", "spss", "csv"],
    })
    assert response.status_code == 200

    with zipfile.ZipFile(io.BytesIO(response.content)) as archive:
        names = set(archive.namelist())
        assert {"household_survey.dta", "household_survey.sav", "household_survey.csv",
                "household_survey_member.dta", "codebook.csv", "README.txt"} <= names

        codebook = archive.read("codebook.csv").decode()
        assert "Household details / Name of village" in codebook
        assert "Male" in codebook
        assert "hh_village" in codebook


def test_export_notes_each_adjustment_once(client):
    response = client.post("/studio/api/export", headers=auth(), json={
        "projectId": 1, "xmlFormId": "household_survey", "formats": ["stata", "spss"],
    })
    with zipfile.ZipFile(io.BytesIO(response.content)) as archive:
        readme = archive.read("README.txt").decode()
    # Both formats mangle names the same way; say so once, not once per format.
    assert readme.count("renamed 'hh-village' to 'hh_village'") == 1


def test_codebook_reports_stable_type_names(client):
    response = client.post("/studio/api/export", headers=auth(), json={
        "projectId": 1, "xmlFormId": "household_survey", "formats": ["csv"],
    })
    with zipfile.ZipFile(io.BytesIO(response.content)) as archive:
        rows = list(csv.DictReader(io.StringIO(archive.read("codebook.csv").decode())))
    types = {r["column"]: r["type"] for r in rows if r["table"] == "household_survey"}
    assert types["hh-village"] == "string"
    assert types["hh-hhsize"] == "integer"
    assert types["SubmissionDate"] == "datetime"

    member = {r["column"]: r["type"] for r in rows if r["table"] == "household_survey_member"}
    # A complete integer column stays an integer; coded selects allow missings.
    assert member["age"] == "integer"
    assert member["sex"] == "numeric"


def test_export_honours_the_selected_language(client):
    response = client.post("/studio/api/export", headers=auth(), json={
        "projectId": 1, "xmlFormId": "household_survey",
        "formats": ["csv"], "language": "Français (fr)",
    })
    with zipfile.ZipFile(io.BytesIO(response.content)) as archive:
        assert "Nom du village" in archive.read("codebook.csv").decode()


def test_export_rejects_an_empty_format_list(client):
    response = client.post("/studio/api/export", headers=auth(), json={
        "projectId": 1, "xmlFormId": "household_survey", "formats": [],
    })
    assert response.status_code == 400


def test_export_is_blocked_for_other_projects(client):
    response = client.post("/studio/api/export", headers=auth("bob-token"), json={
        "projectId": 1, "xmlFormId": "household_survey", "formats": ["csv"],
    })
    assert response.status_code == 403


def test_form_meta_describes_the_form(client):
    meta = client.get("/studio/api/projects/1/forms/household_survey/meta", headers=auth()).json()
    assert meta["title"] == "Household Survey"
    assert meta["repeats"] == ["member"]
    assert "Français (fr)" in meta["languages"]
