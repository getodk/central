from io import BytesIO

from openpyxl import load_workbook

from studio import xlsform
from studio.models import ChoiceList, ChoiceOption, Item, Questionnaire, validate

EN = "English (en)"
FR = "Français (fr)"


def sample():
    return Questionnaire(
        title="Household Survey",
        formId="household_survey",
        languages=[EN, FR],
        defaultLanguage=EN,
        choiceLists=[
            ChoiceList(name="sex_list", options=[
                ChoiceOption(value="1", label={EN: "Male", FR: "Homme"}),
                ChoiceOption(value="2", label={EN: "Female", FR: "Femme"}),
            ]),
            ChoiceList(name="prov", options=[
                ChoiceOption(value="c", label={EN: "Central"}, attributes={"region": "south"}),
            ]),
        ],
        items=[
            Item(kind="group", name="hh", label={EN: "Household"}, children=[
                Item(kind="question", type="text", name="village", label={EN: "Village"}, required=True),
                Item(kind="question", type="integer", name="hhsize", label={EN: "Size"},
                     constraint=". > 0", constraintMessage={EN: "Must be positive"}),
                Item(kind="question", type="select_one", name="province", label={EN: "Province"}, choiceList="prov"),
            ]),
            Item(kind="group", name="member", label={EN: "Members"}, repeat=True,
                 repeatCount="${hhsize}", children=[
                Item(kind="question", type="select_one", name="sex", label={EN: "Sex", FR: "Sexe"}, choiceList="sex_list"),
            ]),
            Item(kind="question", type="calculate", name="total", calculation="count(${member})"),
        ],
    )


def sheets(data):
    workbook = load_workbook(BytesIO(data))
    out = {}
    for name in workbook.sheetnames:
        rows = list(workbook[name].iter_rows(values_only=True))
        header = [c or "" for c in rows[0]]
        out[name] = [dict(zip(header, ["" if c is None else str(c) for c in row])) for row in rows[1:]]
    return out


def test_survey_sheet_structure():
    data = sheets(xlsform.to_workbook(sample()))
    types = [row["type"] for row in data["survey"]]
    assert types == [
        "begin_group", "text", "integer", "select_one prov", "end_group",
        "begin_repeat", "select_one sex_list", "end_repeat", "calculate",
    ]
    repeat = data["survey"][5]
    assert repeat["repeat_count"] == "${hhsize}"


def test_translations_become_per_language_columns():
    data = sheets(xlsform.to_workbook(sample()))
    row = next(r for r in data["survey"] if r["name"] == "sex")
    assert row[f"label::{EN}"] == "Sex"
    assert row[f"label::{FR}"] == "Sexe"
    assert data["settings"][0]["default_language"] == EN


def test_single_language_uses_plain_label_column():
    questionnaire = Questionnaire(
        title="T", formId="t", languages=[EN], defaultLanguage=EN,
        items=[Item(kind="question", type="text", name="q", label={EN: "Q"})],
    )
    data = sheets(xlsform.to_workbook(questionnaire))
    assert "label" in data["survey"][0]
    assert f"label::{EN}" not in data["survey"][0]


def test_choice_attributes_become_extra_columns():
    data = sheets(xlsform.to_workbook(sample()))
    row = next(r for r in data["choices"] if r["name"] == "c")
    assert row["region"] == "south"


def test_unused_columns_are_left_out():
    data = sheets(xlsform.to_workbook(sample()))
    assert "read_only" not in data["survey"][0]
    assert "relevant" not in data["survey"][0]


def test_version_is_generated_when_blank():
    data = sheets(xlsform.to_workbook(sample()))
    assert data["settings"][0]["version"].isdigit()


def test_round_trip_through_a_workbook_is_lossless():
    original = sample()
    original.version = "3"
    restored, warnings = xlsform.from_workbook(xlsform.to_workbook(original))

    assert warnings == []
    assert restored.title == original.title
    assert restored.formId == original.formId
    assert restored.languages == original.languages
    assert restored.defaultLanguage == EN

    flat = {item.name: item for item, _ in restored.walk()}
    assert flat["hh"].kind == "group" and not flat["hh"].repeat
    assert flat["member"].repeat and flat["member"].repeatCount == "${hhsize}"
    assert flat["village"].required is True
    assert flat["hhsize"].constraint == ". > 0"
    assert flat["hhsize"].constraintMessage[EN] == "Must be positive"
    assert flat["sex"].choiceList == "sex_list"
    assert flat["sex"].label[FR] == "Sexe"
    assert flat["total"].calculation == "count(${member})"
    assert [o.value for o in restored.choiceLists[0].options] == ["1", "2"]
    assert restored.choiceLists[1].options[0].attributes["region"] == "south"
    assert validate(restored) == []


def test_unsupported_types_import_as_text_with_a_warning():
    from openpyxl import Workbook

    workbook = Workbook()
    survey = workbook.active
    survey.title = "survey"
    survey.append(["type", "name", "label"])
    survey.append(["select_one_from_file places.csv", "place", "Place"])
    workbook.create_sheet("settings").append(["form_title", "form_id"])
    workbook["settings"].append(["T", "t"])
    buffer = BytesIO()
    workbook.save(buffer)

    restored, warnings = xlsform.from_workbook(buffer.getvalue())
    assert restored.items[0].type == "text"
    assert any("not supported" in w for w in warnings)


def test_unclosed_groups_are_reported():
    from openpyxl import Workbook

    workbook = Workbook()
    survey = workbook.active
    survey.title = "survey"
    survey.append(["type", "name", "label"])
    survey.append(["begin group", "g", "G"])
    survey.append(["text", "q", "Q"])
    buffer = BytesIO()
    workbook.save(buffer)

    restored, warnings = xlsform.from_workbook(buffer.getvalue())
    assert restored.items[0].kind == "group"
    assert restored.items[0].children[0].name == "q"
    assert any("not closed" in w for w in warnings)


def test_validation_catches_common_mistakes():
    broken = Questionnaire(
        title="T", formId="t",
        items=[
            Item(kind="question", type="text", name="1bad", label={EN: "x"}),
            Item(kind="question", type="text", name="dup", label={EN: "x"}),
            Item(kind="question", type="text", name="dup", label={EN: "x"}),
            Item(kind="question", type="select_one", name="s", label={EN: "x"}, choiceList="missing"),
            Item(kind="question", type="calculate", name="c"),
        ],
    )
    messages = [i.message for i in validate(broken) if i.level == "error"]
    assert any("Name must start" in m for m in messages)
    assert any("Duplicate name" in m for m in messages)
    assert any("does not exist" in m for m in messages)
    assert any("calculation is required" in m for m in messages)
