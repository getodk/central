"""Convert between the designer's questionnaire document and XLSForm workbooks.

Generating XLSForm rather than XForms directly means Central converts the form
with the same pyxform it uses for every other form, so a designed form behaves
exactly like an uploaded one.
"""

from __future__ import annotations

import re
import uuid
from datetime import datetime, timezone
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook

from .models import (
    QUESTION_TYPES,
    ChoiceList,
    ChoiceOption,
    Item,
    Questionnaire,
)

_TRUE = {"yes", "true", "1", "y"}

# Columns whose content is per-language.
_TRANSLATABLE = {
    "label": "label",
    "hint": "hint",
    "constraint_message": "constraintMessage",
    "required_message": "requiredMessage",
}


# -- generation ------------------------------------------------------------


def to_workbook(questionnaire: Questionnaire) -> bytes:
    """Render the questionnaire as an XLSForm workbook."""
    multilingual = len(questionnaire.languages) > 1
    languages = questionnaire.languages or [questionnaire.defaultLanguage]

    survey_rows: list[dict[str, str]] = []
    _emit_items(questionnaire.items, survey_rows, questionnaire, languages, multilingual)

    survey_columns = _columns(
        ["type", "name"]
        + _lang_columns("label", languages, multilingual)
        + _lang_columns("hint", languages, multilingual)
        + [
            "required",
            *_lang_columns("required_message", languages, multilingual),
            "relevant",
            "constraint",
            *_lang_columns("constraint_message", languages, multilingual),
            "calculation",
            "default",
            "appearance",
            "read_only",
            "repeat_count",
            "choice_filter",
            "parameters",
        ],
        survey_rows,
        always={"type", "name", *_lang_columns("label", languages, multilingual)},
    )

    choice_rows: list[dict[str, str]] = []
    extra_keys: list[str] = []
    for choice_list in questionnaire.choiceLists:
        for option in choice_list.options:
            row: dict[str, str] = {"list_name": choice_list.name, "name": option.value}
            for column in _lang_columns("label", languages, multilingual):
                lang = column.split("::", 1)[1] if "::" in column else languages[0]
                row[column] = (option.label or {}).get(lang, "")
            for key, value in (option.attributes or {}).items():
                key = _safe_column(key)
                if key and key not in extra_keys:
                    extra_keys.append(key)
                if key:
                    row[key] = value
            choice_rows.append(row)

    choice_columns = (
        ["list_name", "name"]
        + _lang_columns("label", languages, multilingual)
        + extra_keys
    )

    settings_row: dict[str, str] = {
        "form_title": questionnaire.title,
        "form_id": questionnaire.formId,
        "version": questionnaire.version or default_version(),
    }
    if multilingual:
        settings_row["default_language"] = questionnaire.defaultLanguage
    if questionnaire.style:
        settings_row["style"] = questionnaire.style
    if questionnaire.instanceName:
        settings_row["instance_name"] = questionnaire.instanceName
    if questionnaire.publicKey:
        settings_row["public_key"] = questionnaire.publicKey

    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_sheet(workbook, "survey", survey_columns, survey_rows)
    _write_sheet(workbook, "choices", choice_columns, choice_rows)
    _write_sheet(workbook, "settings", list(settings_row), [settings_row])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def default_version() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")


def _lang_columns(base: str, languages: list[str], multilingual: bool) -> list[str]:
    if not multilingual:
        return [base]
    return [f"{base}::{lang}" for lang in languages]


def _safe_column(name: str) -> str:
    return re.sub(r"[^0-9A-Za-z_:() ]+", "", str(name or "")).strip()


def _columns(
    candidates: list[str],
    rows: list[dict[str, str]],
    always: set[str],
) -> list[str]:
    """Keep a column only if something in the sheet uses it."""
    used = {key for row in rows for key, value in row.items() if str(value).strip()}
    return [c for c in candidates if c in always or c in used]


def _write_sheet(
    workbook: Workbook, title: str, columns: list[str], rows: list[dict[str, str]]
) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column, "") for column in columns])


def _emit_items(
    items: list[Item],
    rows: list[dict[str, str]],
    questionnaire: Questionnaire,
    languages: list[str],
    multilingual: bool,
) -> None:
    for item in items:
        if item.kind == "group":
            keyword = "repeat" if item.repeat else "group"
            row = _base_row(item, languages, multilingual)
            row["type"] = f"begin_{keyword}"
            # A section carries its own enabling condition and appearance, which
            # apply to everything nested inside it.
            for source, target in (("relevant", "relevant"), ("appearance", "appearance")):
                value = str(getattr(item, source) or "").strip()
                if value:
                    row[target] = value
            if item.repeat and item.repeatCount.strip():
                row["repeat_count"] = item.repeatCount.strip()
            rows.append(row)
            _emit_items(item.children, rows, questionnaire, languages, multilingual)
            rows.append({"type": f"end_{keyword}", "name": item.name})
            continue

        row = _base_row(item, languages, multilingual)
        row["type"] = _xlsform_type(item)
        if item.required and item.type not in ("note", "calculate", "hidden"):
            row["required"] = "yes"
        if item.readOnly:
            row["read_only"] = "yes"
        for source, target in (
            ("relevant", "relevant"),
            ("constraint", "constraint"),
            ("calculation", "calculation"),
            ("default", "default"),
            ("appearance", "appearance"),
            ("choiceFilter", "choice_filter"),
            ("parameters", "parameters"),
        ):
            value = str(getattr(item, source) or "").strip()
            if value:
                row[target] = value
        rows.append(row)


def _base_row(item: Item, languages: list[str], multilingual: bool) -> dict[str, str]:
    row: dict[str, str] = {"name": item.name}
    for base, attribute in _TRANSLATABLE.items():
        mapping = getattr(item, attribute, None) or {}
        for column in _lang_columns(base, languages, multilingual):
            lang = column.split("::", 1)[1] if "::" in column else languages[0]
            row[column] = mapping.get(lang, "")
    return row


def _xlsform_type(item: Item) -> str:
    spec = QUESTION_TYPES.get(item.type, {})
    if spec.get("choices"):
        return f"{item.type} {item.choiceList}".strip()
    return item.type


# -- import ----------------------------------------------------------------


class ImportError_(Exception):
    """The uploaded workbook is not an XLSForm we can read."""


def from_workbook(data: bytes) -> tuple[Questionnaire, list[str]]:
    """Read an XLSForm workbook back into an editable questionnaire."""
    try:
        workbook = load_workbook(BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:  # openpyxl raises a variety of types
        raise ImportError_(f"could not read the workbook: {exc}") from exc

    sheets = {name.strip().lower(): name for name in workbook.sheetnames}
    if "survey" not in sheets:
        raise ImportError_("the workbook has no 'survey' sheet")

    survey = _read_sheet(workbook[sheets["survey"]])
    choices = _read_sheet(workbook[sheets["choices"]]) if "choices" in sheets else []
    settings = _read_sheet(workbook[sheets["settings"]]) if "settings" in sheets else []

    warnings: list[str] = []
    languages = _languages(survey + choices)
    questionnaire = Questionnaire(languages=languages, defaultLanguage=languages[0])

    if settings:
        row = settings[0]
        questionnaire.title = _get(row, "form_title") or "Imported questionnaire"
        questionnaire.formId = _get(row, "form_id") or "imported_form"
        questionnaire.version = _get(row, "version")
        questionnaire.style = _get(row, "style")
        questionnaire.instanceName = _get(row, "instance_name")
        default_language = _get(row, "default_language")
        if default_language and default_language in languages:
            questionnaire.defaultLanguage = default_language

    questionnaire.choiceLists = _import_choices(choices, languages)
    questionnaire.items, item_warnings = _import_survey(survey, languages)
    warnings.extend(item_warnings)
    return questionnaire, warnings


def _read_sheet(sheet: Any) -> list[dict[str, str]]:
    rows = sheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return []
    columns = [str(c).strip() if c is not None else "" for c in header]
    out: list[dict[str, str]] = []
    for values in rows:
        row = {
            columns[i]: ("" if values[i] is None else str(values[i]).strip())
            for i in range(min(len(columns), len(values)))
            if columns[i]
        }
        if any(row.values()):
            out.append(row)
    return out


def _languages(rows: list[dict[str, str]]) -> list[str]:
    found: list[str] = []
    for row in rows:
        for key in row:
            if key.lower().startswith("label::"):
                lang = key.split("::", 1)[1].strip()
                if lang and lang not in found:
                    found.append(lang)
    return found or ["English (en)"]


def _get(row: dict[str, str], name: str) -> str:
    for key, value in row.items():
        if key.strip().lower() == name:
            return value
    return ""


def _localized(row: dict[str, str], base: str, languages: list[str]) -> dict[str, str]:
    out: dict[str, str] = {}
    plain = _get(row, base)
    if plain:
        out[languages[0]] = plain
    for key, value in row.items():
        lowered = key.strip().lower()
        if lowered.startswith(f"{base}::") and value:
            out[key.split("::", 1)[1].strip()] = value
    return out


def _import_choices(rows: list[dict[str, str]], languages: list[str]) -> list[ChoiceList]:
    known = {"list_name", "list name", "name"}
    lists: dict[str, ChoiceList] = {}
    for row in rows:
        list_name = _get(row, "list_name") or _get(row, "list name")
        if not list_name:
            continue
        option = ChoiceOption(
            value=_get(row, "name"),
            label=_localized(row, "label", languages),
        )
        for key, value in row.items():
            lowered = key.strip().lower()
            if lowered in known or lowered == "label" or lowered.startswith("label::"):
                continue
            if value:
                option.attributes[key.strip()] = value
        lists.setdefault(list_name, ChoiceList(name=list_name)).options.append(option)
    return list(lists.values())


def _import_survey(
    rows: list[dict[str, str]], languages: list[str]
) -> tuple[list[Item], list[str]]:
    root: list[Item] = []
    stack: list[Item] = []
    warnings: list[str] = []

    for index, row in enumerate(rows, start=2):
        raw_type = re.sub(r"\s+", " ", _get(row, "type")).strip()
        if not raw_type:
            continue
        normalised = raw_type.replace("begin ", "begin_").replace("end ", "end_")
        keyword = normalised.split(" ", 1)[0]
        name = _get(row, "name") or f"item_{index}"

        if keyword in ("begin_group", "begin_repeat"):
            item = _common(row, languages)
            item.kind = "group"
            item.repeat = keyword == "begin_repeat"
            item.repeatCount = _get(row, "repeat_count")
            item.relevant = _get(row, "relevant")
            item.appearance = _get(row, "appearance")
            _attach(item, stack, root)
            stack.append(item)
            continue

        if keyword in ("end_group", "end_repeat"):
            if stack:
                stack.pop()
            else:
                warnings.append(f"row {index}: '{raw_type}' without a matching begin")
            continue

        item = _common(row, languages)
        item.kind = "question"
        parts = normalised.split(" ", 1)
        base = parts[0]

        if base in ("select_one", "select_multiple", "rank") and len(parts) > 1:
            item.type = base
            item.choiceList = parts[1].strip()
        elif base in QUESTION_TYPES:
            item.type = base
        else:
            item.type = "text"
            warnings.append(
                f"row {index} ('{name}'): type '{raw_type}' is not supported by the "
                "designer and was imported as text"
            )

        item.required = _get(row, "required").lower() in _TRUE
        item.readOnly = _get(row, "read_only").lower() in _TRUE
        item.relevant = _get(row, "relevant")
        item.constraint = _get(row, "constraint")
        item.calculation = _get(row, "calculation")
        item.default = _get(row, "default")
        item.appearance = _get(row, "appearance")
        item.choiceFilter = _get(row, "choice_filter")
        item.parameters = _get(row, "parameters")
        _attach(item, stack, root)

    if stack:
        warnings.append(
            f"{len(stack)} group(s) were not closed in the workbook; they were "
            "closed at the end of the form"
        )
    return root, warnings


def _common(row: dict[str, str], languages: list[str]) -> Item:
    return Item(
        id=uuid.uuid4().hex[:12],
        name=_get(row, "name"),
        label=_localized(row, "label", languages),
        hint=_localized(row, "hint", languages),
        constraintMessage=_localized(row, "constraint_message", languages),
        requiredMessage=_localized(row, "required_message", languages),
    )


def _attach(item: Item, stack: list[Item], root: list[Item]) -> None:
    (stack[-1].children if stack else root).append(item)
